#!/usr/bin/env python3
"""Quét đơn hàng bưu cục / 3PL remote — KHÔNG đọc danh_sach / export cũ.

Backends: GHN · ViettelPost · SPX(Shopee) · VNPost · (Pancake shipping nếu có token)
Pipeline: nginx /v1/buucuc/scan · panel q:bc_scan · CLI · realtime hook

Owned-only. Cấm dump-login / Acc_all / stealer / demo pad.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

SECRETS = ROOT / "secrets"
OUT = ROOT / "reports" / "telegram-classify"
CACHE = ROOT / "docker" / "nginx-order" / "orders_buucuc_scan_cache.json"
ENV_FILES = (
    SECRETS / "telegram.env",
    SECRETS / "backend_pipes.env",
    SECRETS / "pancake.env",
)

GHN_BASE = "https://online-gateway.ghn.vn/shiip/public-api"
VTP_BASE = "https://partner.viettelpost.vn/v2"


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def load_env() -> dict[str, str]:
    env = dict(os.environ)
    for path in ENV_FILES:
        if not path.is_file():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            t = line.strip()
            if not t or t.startswith("#") or "=" not in t:
                continue
            k, v = t.split("=", 1)
            env.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    try:
        from owned_credentials import env_overlay_from_owned

        env = env_overlay_from_owned(env)
    except Exception:  # noqa: BLE001
        pass
    return env


def _http_json(
    method: str,
    url: str,
    *,
    headers: dict[str, str] | None = None,
    body: dict | list | None = None,
    timeout: float = 25.0,
) -> tuple[int, Any, str]:
    data = None
    hdrs = {"Accept": "application/json", "User-Agent": "buucuc-scan/1.0"}
    if headers:
        hdrs.update(headers)
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    req = urllib.request.Request(url, data=data, headers=hdrs, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            code = resp.getcode() or 0
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="replace")
        code = e.code
    except Exception as e:  # noqa: BLE001
        return 0, None, str(e)[:200]
    try:
        parsed = json.loads(raw) if raw else None
    except json.JSONDecodeError:
        parsed = {"raw": raw[:500]}
    return code, parsed, ""


def _window(days: int) -> tuple[datetime, datetime]:
    """Cửa sổ N ngày lịch (00:00 → now), không rolling 72h — khớp '3 ngày gần nhất'."""
    end = datetime.now(timezone.utc).replace(tzinfo=None)
    n = max(1, int(days))
    start = (end - timedelta(days=n)).replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end


def _norm_order(
    *,
    backend: str,
    buucuc: str,
    order_id: str,
    tracking: str | None,
    status: str | None,
    created_at: str | None,
    shop_id: str | None = None,
    customer_name: str | None = None,
    customer_phone: str | None = None,
    province: str | None = None,
    district: str | None = None,
    ward: str | None = None,
    address: str | None = None,
    cod_amount: float | None = None,
    extra: dict | None = None,
) -> dict[str, Any]:
    oid = str(order_id or tracking or "").strip()
    item: dict[str, Any] = {
        "order_id": oid,
        "tracking_code": tracking or oid,
        "platform": backend,
        "backend": backend,
        "buucuc": buucuc,
        "carrier": backend,
        "status": status,
        "status_shipping": status,
        "status_order": status,
        "shop_id": shop_id,
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "province": province,
        "district": district,
        "ward": ward,
        "address": address,
        "full_address": address,
        "cod_amount": cod_amount,
        "order_created_at": created_at,
        "created_at": created_at,
        "origin": f"buucuc_scan:{backend}",
        "source": "scan_buucuc_orders",
        "channel": "remote_api",
        "shipping": {
            "tracking_code": tracking or oid,
            "status_shipping": status,
            "partner_name": backend,
            "receiver": {
                "name": customer_name,
                "phone": customer_phone,
                "province": province,
                "district": district,
                "ward": ward,
                "address": address,
                "full_address": address,
            },
            "sender": {},
            "warehouse": {},
            "post_office": {"buucuc": buucuc, "carrier": backend, "backend": backend},
        },
    }
    if extra:
        item["raw_keys"] = sorted(extra.keys())[:40]
    return item


# —— GHN ——————————————————————————————————————


def scan_ghn(env: dict[str, str], *, days: int, limit: int) -> dict[str, Any]:
    token = (env.get("GHN_API_TOKEN") or env.get("GHN_TOKEN") or "").strip()
    shop_id = (env.get("GHN_SHOP_ID") or "").strip()
    result: dict[str, Any] = {
        "backend": "GHN",
        "buucuc": "GHN",
        "status": "missing_cred",
        "fetched": 0,
        "orders": [],
        "attempts": [],
        "detail": "",
    }
    if not token:
        result["detail"] = "Thiếu GHN_API_TOKEN — không quét được đơn GHN bưu cục"
        return result

    headers = {"Token": token, "Content-Type": "application/json"}
    if shop_id:
        headers["ShopId"] = shop_id
        headers["ShopID"] = shop_id

    start, end = _window(days)
    # unix seconds commonly accepted by GHN search variants
    from_ts = int(start.timestamp())
    to_ts = int(end.timestamp())

    endpoints = [
        (
            "POST",
            f"{GHN_BASE}/v2/shipping-order/all",
            {
                "payment_type_id": None,
                "required_note": None,
                "from_time": from_ts,
                "to_time": to_ts,
                "offset": 0,
                "limit": min(limit, 200),
            },
        ),
        (
            "POST",
            f"{GHN_BASE}/v2/shipping-order/search",
            {
                "status": [],
                "from_time": from_ts,
                "to_time": to_ts,
                "offset": 0,
                "limit": min(limit, 200),
            },
        ),
        (
            "GET",
            f"{GHN_BASE}/v2/shipping-order/detail",
            None,
        ),
    ]

    orders: list[dict] = []
    seen: set[str] = set()

    for method, url, body in endpoints:
        if len(orders) >= limit:
            break
        # paginate all/search
        if method == "GET":
            code, data, err = _http_json(method, url, headers=headers)
            result["attempts"].append({"url": url, "http": code, "err": err or None})
            continue

        offset = 0
        page_limit = min(200, limit)
        for _ in range(60):  # hard cap pages
            if len(orders) >= limit:
                break
            page_body = dict(body or {})
            page_body["offset"] = offset
            page_body["limit"] = page_limit
            code, data, err = _http_json(method, url, headers=headers, body=page_body)
            result["attempts"].append(
                {"url": url, "offset": offset, "http": code, "err": err or None, "code_field": (data or {}).get("code") if isinstance(data, dict) else None}
            )
            if code == 0:
                result["status"] = "error"
                result["detail"] = err or "network"
                break
            if code in (401, 403):
                result["status"] = "auth_fail"
                result["detail"] = f"GHN auth http={code}"
                return result
            rows = []
            if isinstance(data, dict):
                d = data.get("data")
                if isinstance(d, list):
                    rows = d
                elif isinstance(d, dict):
                    for k in ("orders", "data", "items", "list"):
                        if isinstance(d.get(k), list):
                            rows = d[k]
                            break
                    if not rows and d.get("order_code"):
                        rows = [d]
            if not rows:
                # endpoint not usable / empty
                break
            for row in rows:
                if not isinstance(row, dict):
                    continue
                track = str(row.get("order_code") or row.get("OrderCode") or row.get("sorting_code") or "")
                oid = str(row.get("client_order_code") or track)
                if not oid or oid in seen:
                    continue
                seen.add(oid)
                to_name = row.get("to_name") or row.get("customer_name")
                to_phone = row.get("to_phone") or row.get("customer_phone")
                created = row.get("created_date") or row.get("order_date") or row.get("created_at")
                if isinstance(created, (int, float)):
                    created = datetime.utcfromtimestamp(created).strftime("%Y-%m-%d %H:%M:%S")
                orders.append(
                    _norm_order(
                        backend="GHN",
                        buucuc="GHN",
                        order_id=oid,
                        tracking=track or oid,
                        status=str(row.get("status") or row.get("status_name") or row.get("log") or ""),
                        created_at=str(created) if created else None,
                        shop_id=str(row.get("shop_id") or shop_id or "") or None,
                        customer_name=str(to_name) if to_name else None,
                        customer_phone=str(to_phone) if to_phone else None,
                        address=str(row.get("to_address") or "") or None,
                        province=str(row.get("to_province") or row.get("to_province_name") or "") or None,
                        district=str(row.get("to_district") or row.get("to_district_name") or "") or None,
                        ward=str(row.get("to_ward") or row.get("to_ward_name") or "") or None,
                        cod_amount=float(row["cod_amount"]) if row.get("cod_amount") is not None else None,
                        extra=row,
                    )
                )
                if len(orders) >= limit:
                    break
            if len(rows) < page_limit:
                break
            offset += page_limit
            time.sleep(0.05)

    result["orders"] = orders[:limit]
    result["fetched"] = len(result["orders"])
    if result["fetched"]:
        result["status"] = "ok"
        result["detail"] = f"GHN quét được {result['fetched']} đơn / {days}d"
    elif result["status"] == "missing_cred":
        result["status"] = "empty"
        result["detail"] = "GHN token có nhưng không list được đơn (endpoint/search rỗng hoặc shop_id sai)"
    return result


# —— ViettelPost ——————————————————————————————


def scan_viettelpost(env: dict[str, str], *, days: int, limit: int) -> dict[str, Any]:
    token = (env.get("VIETTELPOST_TOKEN") or "").strip()
    result: dict[str, Any] = {
        "backend": "ViettelPost",
        "buucuc": "ViettelPost",
        "status": "missing_cred",
        "fetched": 0,
        "orders": [],
        "attempts": [],
        "detail": "",
    }
    if not token:
        # thử refresh nếu có user/pass owned
        user = (env.get("VIETTELPOST_USER") or "").strip()
        password = (env.get("VIETTELPOST_PASSWORD") or "").strip()
        if user and password:
            try:
                from access_token_rotate import refresh_viettelpost

                ref = refresh_viettelpost(env)
                result["attempts"].append({"refresh": ref.get("ok"), "verdict": ref.get("verdict")})
                if ref.get("ok"):
                    env = load_env()
                    token = (env.get("VIETTELPOST_TOKEN") or "").strip()
            except Exception as e:  # noqa: BLE001
                result["attempts"].append({"refresh_error": str(e)[:160]})
        if not token:
            result["detail"] = "Thiếu VIETTELPOST_TOKEN (và không refresh được bằng USER/PASSWORD)"
            return result

    headers = {"Token": token, "Content-Type": "application/json"}
    start, end = _window(days)
    # VTP thường dùng dd/MM/yyyy
    frm = start.strftime("%d/%m/%Y %H:%M:%S")
    to = end.strftime("%d/%m/%Y %H:%M:%S")

    # status codes phổ biến VTP: -1 all / 100..500 series
    status_list = [-1, 100, 200, 300, 400, 500, 501, 502, 503, 504]
    orders: list[dict] = []
    seen: set[str] = set()

    endpoints = [
        (f"{VTP_BASE}/order/listOrderByStatus", {"STATUS": -1, "TYPE": 1}),
        (f"{VTP_BASE}/order/listOrderByStatus", {"status": -1}),
        (f"{VTP_BASE}/order/getListOrder", {"FROM": frm, "TO": to}),
        (f"{VTP_BASE}/order/list", {"fromDate": frm, "toDate": to}),
    ]

    for url, base_body in endpoints:
        if len(orders) >= limit:
            break
        for st in status_list if "listOrderByStatus" in url else [None]:
            if len(orders) >= limit:
                break
            body = dict(base_body)
            if st is not None:
                if "STATUS" in body:
                    body["STATUS"] = st
                else:
                    body["status"] = st
            code, data, err = _http_json("POST", url, headers=headers, body=body)
            result["attempts"].append({"url": url, "body_keys": list(body.keys()), "status": st, "http": code, "err": err or None})
            if code in (401, 403):
                result["status"] = "auth_fail"
                result["detail"] = f"VTP auth http={code}"
                return result
            rows = []
            if isinstance(data, dict):
                d = data.get("data")
                if isinstance(d, list):
                    rows = d
                elif isinstance(d, dict):
                    for k in ("orders", "list", "items", "data"):
                        if isinstance(d.get(k), list):
                            rows = d[k]
                            break
            if not isinstance(rows, list):
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                track = str(
                    row.get("ORDER_NUMBER")
                    or row.get("order_number")
                    or row.get("ORDER_NUMBERREFERENCE")
                    or row.get("tracking")
                    or ""
                )
                oid = str(row.get("ORDER_REFERENCE") or row.get("order_reference") or track)
                if not oid or oid in seen:
                    continue
                seen.add(oid)
                created = row.get("ORDER_DATE") or row.get("CREATED_DATE") or row.get("order_date") or row.get("created_at")
                orders.append(
                    _norm_order(
                        backend="ViettelPost",
                        buucuc="ViettelPost",
                        order_id=oid,
                        tracking=track or oid,
                        status=str(row.get("ORDER_STATUS") or row.get("status") or row.get("STATUS_NAME") or ""),
                        created_at=str(created) if created else None,
                        shop_id=(env.get("VIETTELPOST_SHOP_ID") or None),
                        customer_name=str(row.get("RECEIVER_FULLNAME") or row.get("receiver_name") or "") or None,
                        customer_phone=str(row.get("RECEIVER_PHONE") or row.get("receiver_phone") or "") or None,
                        address=str(row.get("RECEIVER_ADDRESS") or row.get("receiver_address") or "") or None,
                        province=str(row.get("RECEIVER_PROVINCE") or row.get("receiver_province") or "") or None,
                        district=str(row.get("RECEIVER_DISTRICT") or "") or None,
                        ward=str(row.get("RECEIVER_WARD") or "") or None,
                        cod_amount=float(row["MONEY_COLLECTION"]) if row.get("MONEY_COLLECTION") is not None else None,
                        extra=row,
                    )
                )
                if len(orders) >= limit:
                    break
            if rows:
                break  # got a working endpoint shape
        if orders:
            break

    result["orders"] = orders[:limit]
    result["fetched"] = len(result["orders"])
    if result["fetched"]:
        result["status"] = "ok"
        result["detail"] = f"VTP quét được {result['fetched']} đơn / {days}d"
    else:
        result["status"] = "empty" if token else "missing_cred"
        result["detail"] = result["detail"] or "VTP token có nhưng listOrder rỗng / endpoint không khớp"
    return result


# —— SPX / Shopee partner (token nếu có) ——————


def scan_spx(env: dict[str, str], *, days: int, limit: int) -> dict[str, Any]:
    token = (env.get("SPX_TOKEN") or env.get("SHOPEE_ACCESS_TOKEN") or "").strip()
    result: dict[str, Any] = {
        "backend": "SPX",
        "buucuc": "SPX",
        "status": "missing_cred",
        "fetched": 0,
        "orders": [],
        "attempts": [],
        "detail": "Thiếu SPX_TOKEN / SHOPEE_ACCESS_TOKEN — không quét remote SPX",
    }
    if not token:
        return result
    # SPX open API varies by partner; probe common logistics list hosts
    shop = (env.get("SPX_SHOP_ID") or env.get("SHOPEE_SHOP_ID") or "").strip()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    urls = [
        "https://spx.vn/api/v1/orders",
        "https://api.spx.vn/api/v1/orders",
    ]
    for url in urls:
        code, data, err = _http_json("GET", url, headers=headers)
        result["attempts"].append({"url": url, "http": code, "err": err or None})
        if code in (401, 403):
            result["status"] = "auth_fail"
            result["detail"] = f"SPX auth http={code}"
            return result
    result["status"] = "unsupported_or_empty"
    result["detail"] = (
        f"SPX token set (shop={shop or '∅'}) nhưng partner list API chưa mở / không trả đơn — "
        "cần endpoint partner cụ thể"
    )
    _ = (days, limit)
    return result


def scan_vnpost(env: dict[str, str], *, days: int, limit: int) -> dict[str, Any]:
    token = (env.get("VNPOST_TOKEN") or "").strip()
    result: dict[str, Any] = {
        "backend": "VNPost",
        "buucuc": "VNPost",
        "status": "missing_cred",
        "fetched": 0,
        "orders": [],
        "attempts": [],
        "detail": "Thiếu VNPOST_TOKEN — không quét remote VNPost",
    }
    if not token:
        return result
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    for url in (
        "https://api.vnpost.vn/api/v1/orders",
        "https://connect-api.vnpost.vn/order/list",
    ):
        code, data, err = _http_json("GET", url, headers=headers)
        result["attempts"].append({"url": url, "http": code, "err": err or None})
        if code in (401, 403):
            result["status"] = "auth_fail"
            result["detail"] = f"VNPost auth http={code}"
            return result
    result["status"] = "unsupported_or_empty"
    result["detail"] = "VNPost token set nhưng list API chưa khớp contract"
    _ = (days, limit)
    return result


def _pancake_tokens(env: dict[str, str]) -> list[dict[str, str]]:
    """Primary + secondary + secrets/pancake_multi_accounts.json (owned)."""
    out: list[dict[str, str]] = []
    seen: set[str] = set()

    def add(token: str, *, name: str = "", uid: str = "") -> None:
        t = (token or "").strip()
        if not t or t in seen:
            return
        seen.add(t)
        out.append({"token": t, "name": name or "", "uid": uid or ""})

    add(env.get("PANCAKE_POS_ACCESS_TOKEN") or "", name=env.get("PANCAKE_USER") or "")
    add(
        env.get("PANCAKE_POS_SECONDARY_ACCESS_TOKEN") or "",
        name=env.get("PANCAKE_POS_SECONDARY_USER") or "secondary",
    )
    raw_multi = (env.get("PANCAKE_POS_ACCESS_TOKENS") or "").strip()
    if raw_multi:
        if raw_multi.startswith("["):
            try:
                arr = json.loads(raw_multi)
                for item in arr:
                    if isinstance(item, str):
                        add(item)
                    elif isinstance(item, dict):
                        add(str(item.get("token") or ""), name=str(item.get("name") or ""))
            except json.JSONDecodeError:
                pass
        else:
            for part in raw_multi.replace(";", ",").split(","):
                add(part.strip())

    multi_path = SECRETS / "pancake_multi_accounts.json"
    if multi_path.is_file():
        try:
            data = json.loads(multi_path.read_text(encoding="utf-8"))
            for acc in data.get("accounts") or []:
                if isinstance(acc, dict):
                    add(
                        str(acc.get("token") or ""),
                        name=str(acc.get("name") or ""),
                        uid=str(acc.get("uid") or ""),
                    )
        except Exception:  # noqa: BLE001
            pass
    return out


def _scan_pancake_one(
    *,
    token: str,
    account_name: str,
    days: int,
    limit: int,
    base: str,
    start: datetime,
    end: datetime,
    seen: set[str],
) -> tuple[list[dict], list[dict], dict[str, Any]]:
    """Quét 1 pos_jwt → shops của đúng account đó (không trộn shop_id account khác)."""
    import requests

    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}",
        "Cookie": f"pos_jwt={token}; pos_locale=vi",
    }
    meta: dict[str, Any] = {"account": account_name, "shops": [], "fetched": 0}
    attempts: list[dict] = []
    try:
        sr = requests.get(f"{base}/shops", headers=headers, timeout=30)
        attempts.append({"account": account_name, "url": f"{base}/shops", "http": sr.status_code})
        shops = (sr.json().get("shops") or []) if sr.ok else []
    except Exception as e:  # noqa: BLE001
        meta["error"] = str(e)[:160]
        return [], attempts, meta

    shop_meta = {
        str(s.get("id")): str(s.get("name") or s.get("shop_name") or "")
        for s in shops
        if s.get("id")
    }
    shop_ids = list(shop_meta.keys())
    meta["shops"] = [{"id": sid, "name": shop_meta[sid]} for sid in shop_ids]
    if not shop_ids:
        meta["detail"] = "token ok nhưng 0 shop"
        return [], attempts, meta

    orders: list[dict] = []
    page_size = 100
    for shop in shop_ids:
        if len(orders) >= limit:
            break
        page = 1
        while len(orders) < limit and page <= 200:
            try:
                rr = requests.get(
                    f"{base}/shops/{shop}/orders",
                    headers=headers,
                    params={"page_number": page, "page_size": page_size},
                    timeout=45,
                )
            except Exception as e:  # noqa: BLE001
                attempts.append({"account": account_name, "shop": shop, "page": page, "error": str(e)[:120]})
                break
            attempts.append({"account": account_name, "shop": shop, "page": page, "http": rr.status_code})
            if not rr.ok:
                break
            rows = rr.json().get("data") or []
            if not rows:
                break
            for row in rows:
                if not isinstance(row, dict):
                    continue
                oid = str(row.get("id") or "")
                if not oid or oid in seen:
                    continue
                created = row.get("inserted_at") or row.get("created_at")
                dt = None
                if created:
                    try:
                        dt = datetime.fromisoformat(str(created).replace("Z", "+00:00")).replace(tzinfo=None)
                    except ValueError:
                        dt = None
                if dt and not (start <= dt <= end):
                    continue
                seen.add(oid)
                sa = row.get("shipping_address") if isinstance(row.get("shipping_address"), dict) else {}
                track = None
                for k in ("tracking_code", "extend_code", "partner_code", "order_shipping_code"):
                    if row.get(k):
                        track = str(row.get(k))
                        break
                partner = row.get("shop_partner_id") or row.get("partner_id")
                wh = row.get("warehouse_info") if isinstance(row.get("warehouse_info"), dict) else {}
                items = row.get("items") if isinstance(row.get("items"), list) else []
                qty = 0
                for it in items:
                    if isinstance(it, dict):
                        try:
                            qty += int(it.get("quantity") or it.get("qty") or 0)
                        except (TypeError, ValueError):
                            pass
                item = _norm_order(
                    backend="Pancake",
                    buucuc=f"Pancake-partner:{partner}" if partner else "Pancake-partner",
                    order_id=oid,
                    tracking=track,
                    status=str(row.get("status_name") or row.get("status") or ""),
                    created_at=str(created) if created else None,
                    shop_id=shop,
                    customer_name=str(sa.get("full_name") or row.get("bill_full_name") or "") or None,
                    customer_phone=str(sa.get("phone_number") or row.get("bill_phone_number") or "") or None,
                    address=str(sa.get("full_address") or sa.get("address") or "") or None,
                    province=str(sa.get("province_name") or "") or None,
                    district=str(sa.get("district_name") or "") or None,
                    ward=str(sa.get("commnue_name") or sa.get("commune_name") or "") or None,
                    cod_amount=float(row["cod"]) if row.get("cod") is not None else None,
                    extra={"partner_id": partner, "warehouse": wh.get("name"), "account": account_name},
                )
                item["warehouse_id"] = row.get("warehouse_id")
                item["kho"] = wh.get("name")
                item["shop_name"] = shop_meta.get(shop) or None
                item["account"] = account_name
                item["custom_id"] = row.get("custom_id") or row.get("display_id")
                item["status_raw"] = row.get("status")
                item["total_price"] = row.get("total_price") or row.get("total") or row.get("cod")
                item["quantity"] = qty or row.get("quantity")
                item["shipping_fee"] = row.get("shipping_fee") or row.get("ship_fee")
                item["partner_id"] = partner
                item["origin"] = "pancake_pos_remote"
                item["source"] = "scan_buucuc_orders"
                orders.append(item)
                if len(orders) >= limit:
                    break
            try:
                last_dt = datetime.fromisoformat(
                    str(rows[-1].get("inserted_at") or "").replace("Z", "+00:00")
                ).replace(tzinfo=None)
            except ValueError:
                last_dt = None
            if last_dt and last_dt < start:
                break
            page += 1

    meta["fetched"] = len(orders)
    return orders, attempts, meta


def scan_pancake_shipping(env: dict[str, str], *, days: int, limit: int) -> dict[str, Any]:
    """Pancake POS — quét MỌI owned token (primary+secondary+multi), mỗi token đúng shop của nó."""
    result: dict[str, Any] = {
        "backend": "Pancake",
        "buucuc": "Pancake-partner",
        "status": "missing_cred",
        "fetched": 0,
        "orders": [],
        "attempts": [],
        "accounts": [],
        "detail": "",
    }
    try:
        import requests  # noqa: F401
    except Exception as e:  # noqa: BLE001
        result["detail"] = f"requests: {e}"
        return result

    tokens = _pancake_tokens(env)
    if not tokens:
        # fallback api_key path
        try:
            from pancake_pos_client import auth_ready, resolve_credentials
        except Exception as e:  # noqa: BLE001
            result["detail"] = f"pancake client: {e}"
            return result
        creds = resolve_credentials()
        if not auth_ready(creds):
            result["detail"] = "Thiếu PANCAKE_POS_API_KEY / Bearer — không quét shipping Pancake"
            return result
        if creds.get("access_token"):
            tokens = [{"token": creds["access_token"], "name": "api", "uid": ""}]
        else:
            result["detail"] = "Chỉ có api_key — cần pos_jwt để quét multi-shop"
            return result

    base = (env.get("PANCAKE_POS_BASE_URL") or "https://pos.pancake.vn/api/v1").rstrip("/")
    start, end = _window(days)
    result["window"] = {"start": start.isoformat(), "end": end.isoformat(), "days": days}

    orders: list[dict] = []
    seen: set[str] = set()
    for acc in tokens:
        if len(orders) >= limit:
            break
        remaining = limit - len(orders)
        chunk, attempts, meta = _scan_pancake_one(
            token=acc["token"],
            account_name=acc.get("name") or "pancake",
            days=days,
            limit=remaining,
            base=base,
            start=start,
            end=end,
            seen=seen,
        )
        result["attempts"].extend(attempts)
        result["accounts"].append(
            {
                "name": meta.get("account"),
                "shops_n": len(meta.get("shops") or []),
                "shops": meta.get("shops"),
                "fetched": meta.get("fetched"),
                "error": meta.get("error"),
                "detail": meta.get("detail"),
            }
        )
        orders.extend(chunk)

    result["orders"] = orders[:limit]
    result["fetched"] = len(result["orders"])
    result["status"] = "ok" if result["fetched"] else "empty"
    acc_bits = ", ".join(
        f"{a.get('name')}={a.get('fetched')}@{a.get('shops_n')}shops" for a in result["accounts"]
    )
    result["detail"] = (
        f"Pancake multi-account quét {result['fetched']} đơn / {days}d "
        f"(accounts={len(result['accounts'])} · window={start.date()}→{end.date()} · {acc_bits})"
    )
    return result


def pipe_into_kho_buucuc(orders: list[dict]) -> dict[str, Any]:
    """Upsert kết quả quét vào pipe kho/bưu cục nếu module có."""
    if not orders:
        return {"ok": True, "upserted": 0, "skipped": True}
    try:
        import sqlite3

        from order_pipe_kho_buucuc_db import (
            BUUCUC_DB,
            PIPE_DB,
            enrich_row,
            ensure_buucuc_mirror_schema,
            ensure_pipe_schema,
            upsert_buucuc_mirror,
            upsert_pipe_order,
        )

        PIPE_DB.parent.mkdir(parents=True, exist_ok=True)
        pipe = sqlite3.connect(str(PIPE_DB))
        mirror = sqlite3.connect(str(BUUCUC_DB))
        ensure_pipe_schema(pipe)
        ensure_buucuc_mirror_schema(mirror)
        n = 0
        for o in orders:
            row = enrich_row(
                {
                    "oms_id": o.get("order_id"),
                    "order_key": o.get("order_id"),
                    "backend": o.get("backend"),
                    "buucuc": o.get("buucuc"),
                    "carrier": o.get("carrier"),
                    "tracking_code": o.get("tracking_code"),
                    "status": o.get("status"),
                    "shop_id": o.get("shop_id"),
                    "shop_name": o.get("shop_name"),
                    "account": o.get("account"),
                    "warehouse_id": o.get("warehouse_id"),
                    "warehouse_name": o.get("kho"),
                    "warehouse_display_name": o.get("kho"),
                    "receiver_name": o.get("customer_name"),
                    "customer_name": o.get("customer_name"),
                    "receiver_phone": o.get("customer_phone"),
                    "customer_phone": o.get("customer_phone"),
                    "province": o.get("province"),
                    "district": o.get("district"),
                    "ward": o.get("ward"),
                    "address_detail": o.get("address"),
                    "full_address": o.get("full_address") or o.get("address"),
                    "cod_amount": o.get("cod_amount"),
                    "created_at": o.get("order_created_at") or o.get("created_at"),
                    "source": "scan_buucuc_orders",
                    "channel": "remote_api",
                    "platform": o.get("platform") or o.get("backend") or "Pancake",
                    "piped_at": utc_now(),
                },
                realtime_new=True,
                pipe_source="buucuc_scan",
            )
            upsert_pipe_order(pipe, row)
            upsert_buucuc_mirror(mirror, row)
            n += 1
        pipe.commit()
        mirror.commit()
        pipe.close()
        mirror.close()
        return {"ok": True, "upserted": n, "pipe_db": str(PIPE_DB), "buucuc_db": str(BUUCUC_DB)}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "upserted": 0, "error": str(e)[:200]}


def build_report(
    *,
    days: int = 3,
    limit: int = 10000,
    backends: list[str] | None = None,
    pipe: bool = True,
    write_cache: bool = True,
    notify: bool = False,
) -> dict[str, Any]:
    env = load_env()
    want = {b.lower() for b in (backends or ["GHN", "ViettelPost", "SPX", "VNPost", "Pancake"])}
    scanners = []
    if "ghn" in want:
        scanners.append(scan_ghn)
    if "viettelpost" in want or "vtp" in want:
        scanners.append(scan_viettelpost)
    if "spx" in want or "shopee" in want:
        scanners.append(scan_spx)
    if "vnpost" in want:
        scanners.append(scan_vnpost)
    if "pancake" in want:
        scanners.append(scan_pancake_shipping)

    backend_reports: list[dict] = []
    all_orders: list[dict] = []
    seen: set[str] = set()
    blockers: list[str] = []

    for fn in scanners:
        remaining = max(0, limit - len(all_orders))
        if remaining <= 0:
            break
        br = fn(env, days=days, limit=remaining)
        backend_reports.append({k: v for k, v in br.items() if k != "orders"})
        if br.get("status") in {"missing_cred", "auth_fail"}:
            blockers.append(f"{br.get('backend')}: {br.get('detail')}")
        for o in br.get("orders") or []:
            key = f"{o.get('backend')}|{o.get('order_id')}|{o.get('tracking_code')}"
            if key in seen:
                continue
            seen.add(key)
            all_orders.append(o)
            if len(all_orders) >= limit:
                break

    pipe_info = pipe_into_kho_buucuc(all_orders) if pipe else {"ok": True, "skipped": True}

    by_buucuc = Counter(str(o.get("buucuc")) for o in all_orders)
    by_backend = Counter(str(o.get("backend")) for o in all_orders)
    by_status = Counter(str(o.get("status")) for o in all_orders)

    report: dict[str, Any] = {
        "ok": True,
        "module": "scan_buucuc_orders",
        "checked_at": utc_now(),
        "request": {
            "days": days,
            "limit": limit,
            "backends": sorted(want),
            "policy": "remote buucuc scan only — không đọc danh_sach/export cũ; cấm demo pad",
        },
        "count": len(all_orders),
        "target": limit,
        "target_met": len(all_orders) >= limit,
        "by_buucuc": dict(by_buucuc),
        "by_backend": dict(by_backend),
        "by_status": dict(by_status),
        "backends": backend_reports,
        "blockers": blockers,
        "pipe": pipe_info,
        "orders": all_orders,
        "verdict": (
            f"Quét bưu cục remote: {len(all_orders)}/{limit} đơn / {days} ngày. "
            + (
                "ĐẠT."
                if len(all_orders) >= limit
                else ("Thiếu token/API — " + "; ".join(blockers[:4]) if blockers else "Endpoints trả rỗng.")
            )
        ),
        "next": [
            "POST /v1/token/set {platform, token, shop_id} qua nginx",
            "python3 scripts/access_token_rotate.py set --platform GHN --token <OWNED> --shop-id <ID>",
            "python3 scripts/access_token_rotate.py set --platform ViettelPost --token <OWNED>",
            "Rồi: python3 scripts/scan_buucuc_orders.py --days 3 --limit 10000 --notify",
        ],
    }

    if write_cache:
        CACHE.parent.mkdir(parents=True, exist_ok=True)
        CACHE.write_text(
            json.dumps(
                {
                    "ok": True,
                    "checked_at": report["checked_at"],
                    "count": report["count"],
                    "total": report["count"],
                    "source": "buucuc_remote_scan",
                    "window_days": days,
                    "by_status_order": dict(by_status),
                    "by_status_shipping": dict(by_status),
                    "orders": all_orders,
                    "blockers": blockers,
                    "verdict": report["verdict"],
                },
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )
        report["cache"] = str(CACHE)

    write_outputs(report)

    excel_info = export_orders_xlsx(all_orders, checked_at=report["checked_at"])
    report["excel"] = excel_info

    if notify:
        try:
            send_telegram_report(env, report)
            report["telegram_notified"] = True
        except Exception as e:  # noqa: BLE001
            report["telegram_notified"] = False
            report["telegram_error"] = str(e)[:160]

    # rewrite after excel/notify so flags persist in summary json
    write_outputs(report)

    return report


EXCEL_HEADERS: tuple[str, ...] = (
    "STT",
    "order_id",
    "custom_id",
    "shop_id",
    "shop_name",
    "platform",
    "status",
    "status_raw",
    "tracking_code",
    "carrier",
    "buucuc",
    "kho",
    "warehouse_id",
    "customer_name",
    "customer_phone",
    "province",
    "district",
    "ward",
    "address",
    "full_address",
    "cod_amount",
    "total_price",
    "quantity",
    "shipping_fee",
    "order_created_at",
    "partner_id",
    "origin",
    "source",
)


def export_orders_xlsx(orders: list[dict], *, checked_at: str) -> dict[str, Any]:
    """Xuất danh_sach_don_hang_chi_tiet.xlsx (nhúng trong pipeline nginx scan)."""
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"openpyxl: {e}"}

    OUT.mkdir(parents=True, exist_ok=True)
    stamp = checked_at.replace(":", "").replace("-", "")[:15]
    primary = OUT / "danh_sach_don_hang_chi_tiet.xlsx"
    stamped = OUT / f"danh_sach_don_hang_chi_tiet_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sach chi tiet"
    ws.append(list(EXCEL_HEADERS))
    for i, o in enumerate(orders, 1):
        ws.append(
            [
                i,
                o.get("order_id"),
                o.get("custom_id"),
                o.get("shop_id"),
                o.get("shop_name"),
                o.get("platform") or o.get("backend") or "Pancake",
                o.get("status") or o.get("status_order"),
                o.get("status_raw") or o.get("status_shipping"),
                o.get("tracking_code"),
                o.get("carrier"),
                o.get("buucuc"),
                o.get("kho"),
                o.get("warehouse_id"),
                o.get("customer_name"),
                o.get("customer_phone"),
                o.get("province"),
                o.get("district"),
                o.get("ward"),
                o.get("address"),
                o.get("full_address") or o.get("address"),
                o.get("cod_amount"),
                o.get("total_price"),
                o.get("quantity"),
                o.get("shipping_fee"),
                o.get("order_created_at") or o.get("created_at"),
                o.get("partner_id"),
                o.get("origin"),
                o.get("source"),
            ]
        )

    ws2 = wb.create_sheet("Tong hop")
    ws2.append(["metric", "value"])
    ws2.append(["checked_at", checked_at])
    ws2.append(["orders", len(orders)])
    by_shop = Counter(str(o.get("shop_id") or "") for o in orders)
    by_status = Counter(str(o.get("status") or "") for o in orders)
    ws2.append(["shops", len(by_shop)])
    for sid, n in by_shop.most_common():
        ws2.append([f"shop:{sid}", n])
    for st, n in by_status.most_common(20):
        ws2.append([f"status:{st}", n])

    ws3 = wb.create_sheet("Theo shop")
    ws3.append(["shop_id", "shop_name", "orders"])
    name_by: dict[str, Any] = {}
    for o in orders:
        sid = str(o.get("shop_id") or "")
        if sid and sid not in name_by and o.get("shop_name"):
            name_by[sid] = o.get("shop_name")
    for sid, n in by_shop.most_common():
        ws3.append([sid, name_by.get(sid), n])

    wb.save(primary)
    wb.save(stamped)
    return {
        "ok": True,
        "path": str(primary),
        "stamped": str(stamped),
        "rows": len(orders),
        "sheets": ["Danh sach chi tiet", "Tong hop", "Theo shop"],
    }


def write_outputs(report: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    slim = {k: v for k, v in report.items() if k != "orders"}
    slim["orders_preview"] = (report.get("orders") or [])[:20]
    jp = OUT / "scan_buucuc_orders.json"
    tp = OUT / "scan_buucuc_orders.txt"
    full = OUT / "scan_buucuc_orders_full.json"
    jp.write_text(json.dumps(slim, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    full.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    tp.write_text(format_text(report) + "\n", encoding="utf-8")
    return {"json": str(jp), "txt": str(tp), "full": str(full)}


def format_text(report: dict[str, Any]) -> str:
    lines = [
        "📦 QUÉT ĐƠN BƯU CỤC (REMOTE)",
        f"Lúc: {report.get('checked_at')}",
        f"Window: {report.get('request', {}).get('days')} ngày gần nhất",
        f"Count: {report.get('count')}/{report.get('target')}",
        f"Verdict: {report.get('verdict')}",
        "",
        "Backends:",
    ]
    for b in report.get("backends") or []:
        lines.append(f"  · {b.get('backend')}: {b.get('status')} — {b.get('detail')} (fetched={b.get('fetched')})")
    if report.get("blockers"):
        lines.append("")
        lines.append("Blockers:")
        for x in report["blockers"]:
            lines.append(f"  - {x}")
    if report.get("by_buucuc"):
        lines.append("")
        lines.append(f"by_buucuc: {report['by_buucuc']}")
    excel = report.get("excel") or {}
    if excel.get("path"):
        lines.append("")
        lines.append(f"Excel: {excel.get('path')} (rows={excel.get('rows')})")
    if report.get("telegram_notified") is not None:
        lines.append(f"Telegram: notified={report.get('telegram_notified')}")
    lines.append("")
    lines.append("Next:")
    for n in report.get("next") or []:
        lines.append(f"  · {n}")
    return "\n".join(lines)


def send_telegram_report(env: dict[str, str], report: dict[str, Any]) -> None:
    token = env.get("TELEGRAM_BOT_TOKEN") or ""
    chat = env.get("TELEGRAM_CHAT_ID") or ""
    if not token or not chat:
        raise RuntimeError("Thiếu TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID")
    text = format_text(report)[:4000]
    payload = json.dumps(
        {"chat_id": chat, "text": text, "disable_web_page_preview": True},
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        resp.read()

    def _send_document(path: Path, caption: str, filename: str) -> None:
        if not path.is_file():
            return
        boundary = "----buucucscan7"
        body = b""

        def part(
            name: str,
            value: bytes | str,
            filename: str | None = None,
            content_type: str | None = None,
        ) -> None:
            nonlocal body
            body += f"--{boundary}\r\n".encode()
            if filename:
                ctype = content_type or "application/octet-stream"
                body += (
                    f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'
                    f"Content-Type: {ctype}\r\n\r\n"
                ).encode()
                body += value if isinstance(value, bytes) else value.encode()
                body += b"\r\n"
            else:
                body += f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode()
                body += str(value).encode() + b"\r\n"

        part("chat_id", chat)
        part("caption", caption[:1000])
        part(
            "document",
            path.read_bytes(),
            filename=filename,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if path.suffix.lower() == ".xlsx"
                else "text/plain; charset=utf-8"
            ),
        )
        body += f"--{boundary}--\r\n".encode()
        req2 = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendDocument",
            data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
            method="POST",
        )
        with urllib.request.urlopen(req2, timeout=120) as resp:
            resp.read()

    txt_path = OUT / "scan_buucuc_orders.txt"
    _send_document(txt_path, f"scan_buucuc · {report.get('count')}/{report.get('target')}", "scan_buucuc_orders.txt")

    excel = report.get("excel") or {}
    xlsx = Path(excel["path"]) if excel.get("path") else OUT / "danh_sach_don_hang_chi_tiet.xlsx"
    _send_document(
        xlsx,
        f"danh_sach_don_hang_chi_tiet · {report.get('count')} đơn / {report.get('request', {}).get('days')}d",
        "danh_sach_don_hang_chi_tiet.xlsx",
    )


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Quét đơn bưu cục remote (GHN/VTP/SPX/VNPost/Pancake)")
    ap.add_argument("--days", type=int, default=3)
    ap.add_argument("--limit", type=int, default=10000)
    ap.add_argument("--backend", action="append", default=None, help="Lặp: GHN, ViettelPost, SPX, VNPost, Pancake")
    ap.add_argument("--no-pipe", action="store_true")
    ap.add_argument("--notify", action="store_true")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args(argv)
    report = build_report(
        days=args.days,
        limit=args.limit,
        backends=args.backend,
        pipe=not args.no_pipe,
        notify=args.notify,
    )
    if args.json:
        slim = {k: v for k, v in report.items() if k != "orders"}
        slim["orders_count"] = report.get("count")
        print(json.dumps(slim, ensure_ascii=False, indent=2, default=str))
    else:
        print(format_text(report))
    return 0 if report.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
