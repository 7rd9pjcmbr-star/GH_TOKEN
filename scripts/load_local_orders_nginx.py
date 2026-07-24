#!/usr/bin/env python3
"""Nạp đơn từ export sở hữu → cache nginx → gọi /orders/local.

Owned-only. Không dump-login.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

INBOX = ROOT / "quarantine" / "telegram"
OUT = ROOT / "reports" / "telegram-classify"
CACHE = ROOT / "docker" / "nginx-order" / "orders_local_cache.json"
DUMP = (
    "acc_all",
    "stealer",
    "ghn_tokens",
    "valid_accounts",
    "results_cookies",
    "vnpost_ok",
    "internal_search",
    "assassin",
)


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def is_dump(name: str) -> bool:
    n = name.lower()
    return any(m in n for m in DUMP)


def load_orders() -> dict:
    orders: list[dict] = []
    sources: Counter[str] = Counter()

    for p in INBOX.glob("orders_detailed*.csv"):
        if is_dump(p.name):
            continue
        with p.open(encoding="utf-8", errors="ignore", newline="") as f:
            for row in csv.DictReader(f):
                orders.append(
                    {
                        "order_id": row.get("remote_id") or row.get("order_key") or row.get("id"),
                        "order_key": row.get("order_key"),
                        "shop_id": row.get("shop_id"),
                        "platform": row.get("platform") or "POS",
                        "status": row.get("status_normalized") or row.get("status_raw"),
                        "customer_name": row.get("customer_name"),
                        "customer_phone": row.get("customer_phone"),
                        "source": row.get("source") or p.name,
                        "file": p.name,
                        "backend": "local-csv",
                    }
                )
                sources[p.name] += 1

    for p in INBOX.glob("orders_detailed*.json"):
        if is_dump(p.name):
            continue
        raw = json.loads(p.read_text(encoding="utf-8", errors="ignore"))
        items = raw if isinstance(raw, list) else []
        if isinstance(raw, dict):
            for k in ("orders", "data", "items", "results"):
                if isinstance(raw.get(k), list):
                    items = raw[k]
                    break
        n = 0
        for it in items:
            if not isinstance(it, dict):
                continue
            status = it.get("status") or it.get("status_name") or it.get("status_normalized")
            if isinstance(status, dict):
                status = status.get("name") or status.get("code") or str(status)
            phone = it.get("bill_phone_number") or it.get("customer_phone")
            sa = it.get("shipping_address")
            if not phone and isinstance(sa, dict):
                phone = sa.get("phone")
            tracking = None
            for k in ("tracking_code", "extend_code", "partner_code", "order_shipping_code"):
                if it.get(k):
                    tracking = it.get(k)
                    break
            orders.append(
                {
                    "order_id": str(it.get("id") or it.get("order_id") or it.get("display_id") or ""),
                    "shop_id": str(it.get("shop_id") or "") or None,
                    "platform": "Pancake/POS",
                    "status": str(status) if status is not None else None,
                    "tracking_code": tracking,
                    "customer_name": it.get("bill_full_name") or it.get("customer_name"),
                    "customer_phone": phone,
                    "source": p.name,
                    "file": p.name,
                    "backend": "local-json",
                    "page_id": it.get("page_id"),
                    "warehouse_id": it.get("warehouse_id"),
                }
            )
            n += 1
        sources[p.name] += n

    try:
        import openpyxl
    except ImportError:
        openpyxl = None

    thanh = INBOX / "thanhcoong.xlsx"
    if thanh.is_file() and openpyxl:
        wb = openpyxl.load_workbook(thanh, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h or "") for h in rows[0]]
        n = 0

        def col(r, name):
            if name not in hdr:
                return None
            i = hdr.index(name)
            return r[i] if i < len(r) else None

        for r in rows[1:]:
            if not r:
                continue
            track = col(r, "Tracking No.")
            if not track or str(track).startswith("Tracking"):
                continue
            orders.append(
                {
                    "order_id": str(track),
                    "tracking_code": str(track),
                    "shop_id": str(col(r, "Account ID") or ""),
                    "platform": "SPX",
                    "status": str(col(r, "Tracking Status") or ""),
                    "customer_name": str(col(r, "Receiver Name") or ""),
                    "customer_phone": str(col(r, "Receiver Phone Number") or ""),
                    "province": str(col(r, "Receiver Province") or ""),
                    "source": "thanhcoong.xlsx",
                    "file": "thanhcoong.xlsx",
                    "backend": "SPX-local",
                }
            )
            n += 1
        wb.close()
        sources["thanhcoong.xlsx"] = n

    ds = INBOX / "danh_sach_don_hang_20260512.xlsx"
    if ds.is_file() and openpyxl:
        wb = openpyxl.load_workbook(ds, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h or "") for h in rows[0]]
        n = 0
        for r in rows[1:]:
            if not r:
                continue
            row = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
            oid = row.get("order_id")
            if not oid:
                continue
            orders.append(
                {
                    "order_id": str(oid),
                    "platform": str(row.get("platform") or ""),
                    "status": str(row.get("status") or ""),
                    "customer_name": str(row.get("customer_name") or ""),
                    "source": "danh_sach_don_hang_20260512.xlsx",
                    "file": "danh_sach_don_hang_20260512.xlsx",
                    "backend": "local-xlsx",
                }
            )
            n += 1
        wb.close()
        sources["danh_sach_don_hang_20260512.xlsx"] = n

    seen = set()
    uniq = []
    for o in orders:
        key = (str(o.get("order_id")), o.get("file"))
        if key in seen:
            continue
        seen.add(key)
        uniq.append(o)

    stamp = utc_now()
    return {
        "ok": True,
        "checked_at": stamp,
        "source": "owned_local_exports",
        "count": len(uniq),
        "by_file": dict(sources),
        "by_platform": dict(Counter(o.get("platform") or "?" for o in uniq)),
        "by_status": dict(Counter((o.get("status") or "?")[:40] for o in uniq).most_common(20)),
        "by_shop": dict(Counter(str(o.get("shop_id") or "") for o in uniq).most_common(10)),
        "orders": uniq,
        "policy": {"owned_only": True, "no_dump_login": True},
    }


def write_outputs(payload: dict) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    CACHE.parent.mkdir(parents=True, exist_ok=True)
    (OUT / "orders_local_owned.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    lines = [
        f"📦 ĐƠN LOCAL SỞ HỮU · {payload.get('checked_at')}",
        f"Tổng: {payload.get('count')}",
        f"by_file={payload.get('by_file')}",
        f"by_platform={payload.get('by_platform')}",
        f"by_shop={payload.get('by_shop')}",
        "",
    ]
    for o in (payload.get("orders") or [])[:40]:
        lines.append(
            f"· {o.get('order_id')} · shop={o.get('shop_id')} · {o.get('status')} · "
            f"{o.get('platform')} · {o.get('tracking_code') or ''} · {o.get('file')}"
        )
    extra = max(0, int(payload.get("count") or 0) - 40)
    if extra:
        lines.append(f"… +{extra} đơn nữa")
    lines += [
        "",
        "Nginx: GET http://127.0.0.1:18080/orders/local?limit=100",
        "       GET http://127.0.0.1:18080/v1/orders/local?shop_id=1530618",
    ]
    (OUT / "orders_local_owned.txt").write_text("\n".join(lines), encoding="utf-8")
    cache = {
        "ok": True,
        "checked_at": payload.get("checked_at"),
        "count": payload.get("count"),
        "total": payload.get("count"),
        "orders": payload.get("orders"),
        "source": "owned_local_exports",
        "by_platform": payload.get("by_platform"),
        "by_shop": payload.get("by_shop"),
    }
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, default=str), encoding="utf-8")


def serve_via_nginx(*, keep: bool = True) -> dict:
    from nginx_order_embed import NginxOrderEmbed

    mod = NginxOrderEmbed(auto_stop=not keep)
    up = mod.ensure_up()
    if not up.get("ok"):
        return {"ok": False, "error": "nginx not up", "start": up}
    res = mod.call_json("/orders/local?limit=20", method="GET", ensure=False)
    return {
        "ok": bool(res.get("ok")),
        "via_nginx": True,
        "embedded": res.get("embedded"),
        "http": res.get("http"),
        "payload_count": (res.get("payload") or {}).get("count"),
        "payload_total": (res.get("payload") or {}).get("total"),
        "sample": ((res.get("payload") or {}).get("orders") or [])[:10],
    }


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Nạp đơn local sở hữu → nginx /orders/local")
    ap.add_argument("--no-nginx", action="store_true")
    ap.add_argument("--stop", action="store_true", help="tắt nginx sau khi gọi")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args(argv)

    payload = load_orders()
    write_outputs(payload)
    ngx = None
    if not args.no_nginx:
        ngx = serve_via_nginx(keep=not args.stop)
        if args.stop:
            from nginx_order_embed import NginxOrderEmbed

            NginxOrderEmbed().stop()

    report = {
        "ok": True,
        "checked_at": payload.get("checked_at"),
        "count": payload.get("count"),
        "by_platform": payload.get("by_platform"),
        "by_shop": payload.get("by_shop"),
        "by_file": payload.get("by_file"),
        "nginx": ngx,
        "verdict": f"✅ Đã lấy {payload.get('count')} đơn local sở hữu"
        + (f" · nginx total={((ngx or {}).get('payload_total'))}" if ngx else ""),
        "endpoints": [
            "http://127.0.0.1:18080/orders/local?limit=100",
            "http://127.0.0.1:18080/v1/orders/local?shop_id=1530618",
            "http://127.0.0.1:18080/v1/orders/local?platform=SPX&limit=50",
        ],
    }
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    else:
        print(report["verdict"])
        print("by_platform", report["by_platform"])
        print("by_shop", report["by_shop"])
        print("by_file", report["by_file"])
        if ngx:
            print("nginx", ngx.get("http"), "total", ngx.get("payload_total"), "upstream", (ngx.get("embedded") or {}).get("$upstream_addr"))
            for o in ngx.get("sample") or []:
                print(f"  · {o.get('order_id')} · {o.get('shop_id')} · {o.get('status')} · {o.get('platform')}")
        for ep in report["endpoints"]:
            print("·", ep)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
