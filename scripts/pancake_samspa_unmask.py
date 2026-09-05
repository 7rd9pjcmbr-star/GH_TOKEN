#!/usr/bin/env python3
"""SamSpa 1530618: ghép đơn API (SĐT mask) với nguồn Excel/inbox (SĐT đầy đủ).

Pancake API luôn trả bill_phone_number có **** — không bẻ mask qua API.
Chỉ backfill từ file export sở hữu hoặc carrier khi có mã vận đơn.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

INBOX = ROOT / "quarantine" / "telegram"
REPORTS = ROOT / "reports" / "telegram-classify"
SAMSPA_SHOP = "1530618"


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def is_masked_phone(raw: str) -> bool:
    from fix_order_phones import is_masked

    return is_masked(str(raw or ""))


def phone_index_from_owned() -> dict[str, str]:
    """Map order_key / remote_id / tracking / custom_id → SĐT đầy đủ (owned files)."""
    idx: dict[str, str] = {}

    def add(keys: list[str], phone: str) -> None:
        if not phone or is_masked_phone(phone):
            return
        for k in keys:
            k = (k or "").strip()
            if k:
                idx[k] = phone

    paths: list[Path] = [REPORTS / "KET_QUA_DON_CHIET_TIET.csv"]
    paths.extend(sorted(INBOX.glob("orders_detailed_*.csv")))
    paths.extend(sorted(INBOX.glob("orders*.xlsx")))
    paths.extend(sorted(INBOX.glob("*orders*.xlsx")))

    for p in paths:
        if not p.is_file():
            continue
        if p.suffix.lower() == ".csv":
            try:
                with p.open(encoding="utf-8", errors="replace", newline="") as fh:
                    for row in csv.DictReader(fh):
                        ph = (row.get("customer_phone") or row.get("SĐT") or row.get("SDT") or "").strip()
                        add(
                            [
                                row.get("order_key") or "",
                                row.get("remote_id") or "",
                                row.get("tracking_code") or "",
                                row.get("custom_id") or "",
                            ],
                            ph,
                        )
            except OSError:
                continue
        elif p.suffix.lower() == ".xlsx":
            try:
                import openpyxl

                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                for sheet in wb.sheetnames[:3]:
                    ws = wb[sheet]
                    rows = ws.iter_rows(values_only=True)
                    header = [str(c or "").strip() for c in next(rows, [])]
                    if not header:
                        continue
                    col = {h: i for i, h in enumerate(header)}
                    phone_i = next(
                        (col[h] for h in header if h.lower() in {"customer_phone", "sđt", "sdt", "phone"}),
                        None,
                    )
                    if phone_i is None:
                        continue
                    id_cols = [
                        col.get(h)
                        for h in header
                        if h.lower()
                        in {
                            "order_key",
                            "remote_id",
                            "tracking_code",
                            "custom_id",
                            "mã đơn hàng",
                            "ma don hang",
                            "id",
                        }
                        and col.get(h) is not None
                    ]
                    for row in rows:
                        if not row or phone_i >= len(row):
                            continue
                        ph = str(row[phone_i] or "").strip()
                        keys = [str(row[i] or "").strip() for i in id_cols if i < len(row)]
                        add(keys, ph)
                wb.close()
            except Exception:  # noqa: BLE001
                continue
    return idx


def fetch_samspa_api_rows(*, days: int, limit: int) -> list[dict]:
    from datetime import datetime as dt, timedelta

    import requests

    from export_orders_detailed import load_env, scan_order_to_row

    env = load_env()
    token = env.get("PANCAKE_POS_ACCESS_TOKEN") or env.get("PANCAKE_POS_API_KEY") or ""
    if not token:
        return []
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Cookie": f"pos_jwt={token}; pos_locale=vi",
    }
    base = "https://pos.pancake.vn/api/v1"
    start = dt.utcnow() - timedelta(days=days)
    end = dt.utcnow()
    orders: list[dict] = []
    page = 1
    while len(orders) < limit and page <= 200:
        rr = requests.get(
            f"{base}/shops/{SAMSPA_SHOP}/orders",
            headers=headers,
            params={"page_number": page, "page_size": 100},
            timeout=45,
        )
        if not rr.ok:
            break
        rows = rr.json().get("data") or []
        if not rows:
            break
        for row in rows:
            if not isinstance(row, dict):
                continue
            created = row.get("inserted_at") or row.get("created_at")
            if created:
                try:
                    dt_created = dt.fromisoformat(str(created).replace("Z", "+00:00")).replace(tzinfo=None)
                    if not (start <= dt_created <= end):
                        continue
                except ValueError:
                    pass
            sa = row.get("shipping_address") if isinstance(row.get("shipping_address"), dict) else {}
            track = next((str(row.get(k)) for k in ("tracking_code", "extend_code", "partner_code") if row.get(k)), "")
            orders.append(
                scan_order_to_row(
                    {
                        "order_id": str(row.get("id") or ""),
                        "custom_id": row.get("custom_id") or row.get("display_id"),
                        "shop_id": SAMSPA_SHOP,
                        "shop_name": "SamSpa Shop",
                        "customer_name": sa.get("full_name") or row.get("bill_full_name"),
                        "customer_phone": sa.get("phone_number") or row.get("bill_phone_number"),
                        "tracking_code": track,
                        "status": row.get("status_name") or row.get("status"),
                        "status_raw": row.get("status"),
                        "province": sa.get("province_name"),
                        "district": sa.get("district_name"),
                        "ward": sa.get("commnue_name") or sa.get("commune_name"),
                        "address": sa.get("full_address") or sa.get("address"),
                        "created_at": created,
                        "platform": "Pancake/POS",
                        "backend": "Pancake",
                        "carrier": "Pancake",
                        "source": "pancake_api_samspa",
                    },
                    "pancake_api_samspa",
                )
            )
            if len(orders) >= limit:
                break
        page += 1
    return orders


def backfill_rows(rows: list[dict], idx: dict[str, str]) -> tuple[list[dict], dict]:
    stats = Counter()
    out: list[dict] = []
    for r in rows:
        row = dict(r)
        raw = str(row.get("customer_phone") or "")
        if is_masked_phone(raw) or not raw:
            keys = [
                row.get("order_key") or "",
                row.get("remote_id") or "",
                row.get("tracking_code") or "",
                row.get("custom_id") or "",
            ]
            found = None
            for k in keys:
                k = str(k).strip()
                if k and k in idx:
                    found = idx[k]
                    break
            if found:
                row["customer_phone"] = found
                row["phone_status"] = "backfilled_owned"
                row["phone_source"] = "owned_export"
                stats["backfilled"] += 1
            else:
                row["phone_status"] = "masked_api"
                row["phone_source"] = "pancake_api"
                stats["still_masked"] += 1
        else:
            row["phone_status"] = "ok"
            row["phone_source"] = row.get("source") or "owned"
            stats["ok"] += 1
        out.append(row)
    return out, dict(stats)


def write_samspa_outputs(rows: list[dict], stats: dict, idx_n: int) -> dict[str, Path]:
    REPORTS.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    csv_path = REPORTS / f"SAMSPA_1530618_DAY_DU_{ts}.csv"
    json_path = REPORTS / f"SAMSPA_1530618_DAY_DU_{ts}.json"
    summary_path = REPORTS / "SAMSPA_UNMASK_SUMMARY.txt"

    fields = list(rows[0].keys()) if rows else ["order_key", "customer_phone", "phone_status"]
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    summary_path.write_text(
        "\n".join(
            [
                f"SAMSPA UNMASK · {utc_now()}",
                f"shop_id={SAMSPA_SHOP}",
                f"phone_index_keys={idx_n}",
                f"rows={len(rows)}",
                f"stats={stats}",
                "",
                "Pancake API luôn mask SĐT (****). File này ưu tiên SĐT từ Excel/inbox sở hữu.",
                "Đơn còn masked_api: export Excel từ pos.pancake.vn → gửi bot Telegram.",
                "",
                f"csv={csv_path}",
                f"json={json_path}",
            ]
        ),
        encoding="utf-8",
    )
    return {"csv": csv_path, "json": json_path, "summary": summary_path}


def build_report(*, days: int = 7, limit: int = 5000, api: bool = True) -> dict:
    idx = phone_index_from_owned()
    rows: list[dict] = []

    if api:
        rows.extend(fetch_samspa_api_rows(days=days, limit=limit))

    # Owned inbox rows tagged SamSpa or from samspa-named exports
    from export_orders_detailed import CSV_FIELDS, load_inbox_rows

    for r in load_inbox_rows():
        sid = str(r.get("shop_id") or "")
        sname = str(r.get("shop_name") or "").lower()
        if sid == SAMSPA_SHOP or "samspa" in sname or "sam spa" in sname:
            rows.append(r)

    # Dedupe by order key — prefer full phone
    merged: dict[str, dict] = {}
    for r in rows:
        key = str(r.get("order_key") or r.get("remote_id") or "")
        if not key:
            continue
        prev = merged.get(key)
        if prev is None:
            merged[key] = r
            continue
        pick = r
        if is_masked_phone(r.get("customer_phone") or "") and not is_masked_phone(prev.get("customer_phone") or ""):
            pick = prev
        elif is_masked_phone(prev.get("customer_phone") or "") and not is_masked_phone(r.get("customer_phone") or ""):
            pick = r
        else:
            pick = {**prev, **{k: v for k, v in r.items() if v}}
        merged[key] = pick

    rows = list(merged.values())
    filled, stats = backfill_rows(rows, idx)
    paths = write_samspa_outputs(filled, stats, len(idx)) if filled else {}

    return {
        "ok": True,
        "module": "pancake_samspa_unmask",
        "checked_at": utc_now(),
        "shop_id": SAMSPA_SHOP,
        "phone_index": len(idx),
        "rows": len(filled),
        "stats": stats,
        "paths": {k: str(v) for k, v in paths.items()},
        "verdict": (
            f"SamSpa {len(filled)} đơn · backfill={stats.get('backfilled', 0)} · "
            f"masked_api={stats.get('still_masked', 0)} · ok={stats.get('ok', 0)}"
        ),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="SamSpa: backfill SĐT từ export owned")
    ap.add_argument("--days", type=int, default=7)
    ap.add_argument("--limit", type=int, default=5000)
    ap.add_argument("--no-api", action="store_true")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    rep = build_report(days=args.days, limit=args.limit, api=not args.no_api)
    if args.json:
        print(json.dumps(rep, ensure_ascii=False, indent=2))
    else:
        print(rep["verdict"])
        for k, v in rep.get("paths", {}).items():
            print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
